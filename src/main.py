"""CLI entry point — Indian Equity Long-Term Investor."""
from __future__ import annotations

import asyncio
import re
from datetime import date
from pathlib import Path

import click
from rich.console import Console
from rich.table import Table

from src.agent.pipeline import InvestmentPipeline
from src.config import settings
from src.logging_config import configure_logging
from src.models import AnalysisState
from src.portfolio.tracker import PortfolioTracker, add_one_year

_TICKER_RE = re.compile(r"^[A-Z0-9&\-\.]{1,20}$")


def _validate_ticker(ticker: str) -> str:
    """Normalise and validate a ticker symbol. Raises click.BadParameter on invalid input."""
    normalised = ticker.upper().strip()
    if not _TICKER_RE.match(normalised):
        raise click.BadParameter(
            f"'{ticker}' is not a valid NSE ticker. "
            "Expected 1–20 uppercase alphanumeric characters (e.g. RELIANCE, M&M, BAJAJ-AUTO)."
        )
    return normalised

console = Console()


# ---------------------------------------------------------------------------
# CLI group
# ---------------------------------------------------------------------------


@click.group()
@click.version_option("1.0.0")
def cli() -> None:
    """Indian Equity Long-Term Investor — AI-powered stock analysis.

    Uses a 9-step research pipeline backed by Claude AI to evaluate
    long-term investment opportunities on NSE/BSE.
    """
    configure_logging(log_level=settings.log_level, log_format=settings.log_format)


# ---------------------------------------------------------------------------
# analyze command
# ---------------------------------------------------------------------------


@cli.command()
@click.argument("ticker")
@click.option(
    "--save/--no-save",
    default=True,
    help="Save report to analysis/reports/ (default: True)",
)
def analyze(ticker: str, save: bool) -> None:
    """Run full 9-step analysis on a stock TICKER.

    Example: investor analyze RELIANCE
    """
    ticker = _validate_ticker(ticker)

    async def _run() -> AnalysisState:
        pipeline = InvestmentPipeline()
        return await pipeline.analyze(ticker)

    with console.status(f"[bold green]Analysing {ticker.upper()}...[/bold green]"):
        state = asyncio.run(_run())

    if state.formatted_output:
        console.print(state.formatted_output)
    else:
        console.print(f"[yellow]No formatted output generated for {ticker}[/yellow]")

    if save and state.formatted_output:
        _save_report(ticker, state)


def _save_report(ticker: str, state: AnalysisState) -> None:
    """Save the formatted report to analysis/reports/."""
    reports_dir = Path("analysis") / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    report_path = reports_dir / f"{ticker.upper()}_{today}.md"
    report_path.write_text(state.formatted_output or "", encoding="utf-8")
    console.print(f"[dim]Report saved → {report_path}[/dim]")


# ---------------------------------------------------------------------------
# portfolio command
# ---------------------------------------------------------------------------


@cli.command()
@click.option(
    "--user",
    "user_id",
    default=None,
    envvar="INVESTOR_USER",
    help="Portfolio user ID (default: settings.investor_user / INVESTOR_USER env var).",
)
def portfolio(user_id: str | None) -> None:
    """Show current portfolio holdings and summary."""
    tracker = PortfolioTracker(user_id=user_id)
    holdings = asyncio.run(tracker.get_holdings())

    if not holdings:
        console.print(f"[yellow]No holdings found for user '{tracker.user_id}'.[/yellow]")
        return

    table = Table(title="Portfolio Holdings", show_header=True, header_style="bold cyan")
    table.add_column("Ticker", style="bold")
    table.add_column("Company")
    table.add_column("Avg Cost", justify="right")
    table.add_column("Qty", justify="right")
    table.add_column("Purchase Date")
    table.add_column("Allocation %", justify="right")

    total_alloc = 0.0
    for h in holdings:
        table.add_row(
            h["ticker"],
            h.get("company_name", ""),
            f"₹{h['avg_cost']:.2f}",
            str(h["quantity"]),
            h.get("purchase_date", ""),
            f"{h['allocation_pct']:.1f}%",
        )
        total_alloc += h["allocation_pct"]

    console.print(table)
    console.print(f"\n[bold]Total Allocation: {total_alloc:.1f}%[/bold]")


# ---------------------------------------------------------------------------
# add-trade command
# ---------------------------------------------------------------------------


@cli.command("add-trade")
@click.argument("ticker")
@click.argument("action", type=click.Choice(["BUY", "SELL"], case_sensitive=False))
@click.argument("qty", type=int)
@click.argument("price", type=str)  # str so we can strip commas before converting
@click.option(
    "--date",
    "txn_date",
    default=None,
    help="Trade date in YYYY-MM-DD format (default: today).",
)
@click.option("--company", default="", help="Company name (BUY only).")
@click.option("--allocation", "allocation_pct", default=0.0, type=float, help="Allocation % of portfolio (BUY only).")
@click.option("--notes", default="", help="Optional notes recorded in the transaction log.")
@click.option(
    "--user",
    "user_id",
    default=None,
    envvar="INVESTOR_USER",
    help="Portfolio user ID.",
)
def add_trade(
    ticker: str,
    action: str,
    qty: int,
    price: str,
    txn_date: str | None,
    company: str,
    allocation_pct: float,
    notes: str,
    user_id: str | None,
) -> None:
    """Record a BUY or SELL transaction in the portfolio database.

    \b
    Examples:
      investor add-trade RELIANCE BUY 10 2850.50
      investor add-trade INFY SELL 5 1920.00 --date 2026-05-15 --notes "partial exit"
      investor add-trade HDFCBANK BUY 20 1650.00 --company "HDFC Bank" --allocation 5.0
      investor add-trade RELIANCE BUY 10 2850.50 --user rm
    """
    ticker = _validate_ticker(ticker)
    action = action.upper()

    try:
        price_f = float(price.replace(",", ""))
    except ValueError:
        raise click.BadParameter(
            f"'{price}' is not a valid price. Use a number e.g. 1062.00 or 1,062.00",
            param_hint="price",
        )

    try:
        trade_date = date.fromisoformat(txn_date) if txn_date else date.today()
    except ValueError:
        raise click.BadParameter(
            f"'{txn_date}' is not a valid date. Use YYYY-MM-DD format.",
            param_hint="--date",
        )

    if qty <= 0:
        raise click.BadParameter("QTY must be a positive integer.", param_hint="qty")
    if price_f <= 0:
        raise click.BadParameter("PRICE must be positive.", param_hint="price")

    tracker = PortfolioTracker(user_id=user_id)

    async def _record() -> dict | None:
        if action == "BUY":
            await tracker.add_transaction(
                ticker=ticker, action=action, price=price_f,
                quantity=qty, txn_date=trade_date, notes=notes,
            )
            await tracker.add_holding(
                ticker=ticker, avg_cost=price_f, quantity=qty,
                purchase_date=trade_date, allocation_pct=allocation_pct,
                company_name=company or ticker,
            )
            await tracker.add_tax_entry(
                ticker=ticker, purchase_date=trade_date,
                ltcg_date=add_one_year(trade_date), avg_cost=price_f,
            )
            return None
        # SELL — consume purchase lots FIFO; record_sell logs the transaction
        # only after the lots are successfully consumed.
        return await tracker.record_sell(
            ticker=ticker, price=price_f, quantity=qty,
            txn_date=trade_date, notes=notes,
        )

    try:
        sell_result = asyncio.run(_record())
    except ValueError as exc:
        # Oversell: nothing was recorded — surface a clean CLI error.
        raise click.ClickException(str(exc)) from exc

    console.print(
        f"[green]✓[/green] Transaction logged: {action} {qty} × {ticker} @ ₹{price_f:.2f} on {trade_date}"
        f"  [dim](user: {tracker.user_id})[/dim]"
    )
    if action == "BUY":
        console.print("[green]✓[/green] Holding added to portfolio DB")
        console.print(
            f"[green]✓[/green] Tax tracker updated — LTCG eligible from {add_one_year(trade_date)} "
            f"(gains > ₹1.25L taxed at 12.5% after that date)"
        )
    if action == "SELL" and sell_result:
        gain_colour = "green" if sell_result["realized_gain"] >= 0 else "red"
        console.print(
            f"[green]✓[/green] {qty} shares consumed FIFO across "
            f"{len(sell_result['lots'])} lot(s); {sell_result['remaining_qty']} still held"
        )
        for lot in sell_result["lots"]:
            tax_label = "LTCG" if lot["is_ltcg"] else "STCG"
            lot_colour = "green" if lot["gain"] >= 0 else "red"
            console.print(
                f"  Lot {lot['purchase_date']} @ ₹{lot['avg_cost']:.2f} × "
                f"{lot['quantity_consumed']}: [{lot_colour}]₹{lot['gain']:+,.2f}[/{lot_colour}] ({tax_label})"
            )
        console.print(
            f"Realized P&L: [{gain_colour}]₹{sell_result['realized_gain']:+,.2f}[/{gain_colour}]  "
            f"(LTCG ₹{sell_result['ltcg_gain']:+,.2f} — 12.5% over ₹1.25L/yr | "
            f"STCG ₹{sell_result['stcg_gain']:+,.2f} — 20%)"
        )


# ---------------------------------------------------------------------------
# watchlist command
# ---------------------------------------------------------------------------


@cli.command()
@click.argument("ticker")
@click.argument("tier", type=click.Choice(["1", "2", "3"]))
@click.option("--reason", default="", help="Reason for watchlist addition")
def watchlist(ticker: str, tier: str, reason: str) -> None:
    """Show how to add TICKER to the watchlist.

    Watchlist entries are now managed automatically by the analysis pipeline
    and persisted to the SQLite database (investor.db).

    Run a full analysis instead:  investor analyze TICKER
    View watchlist alerts:         investor watchlist-alerts
    View all watchlist tickers:    investor db-recommendations --type WATCHLIST
    """
    console.print(
        f"[yellow]ℹ Watchlist is now managed automatically by the analysis pipeline.[/yellow]\n"
        f"Run [bold]investor analyze {ticker.upper()}[/bold] to evaluate and add to watchlist.\n"
        f"View current watchlist: [bold]investor db-recommendations --type WATCHLIST[/bold]\n"
        f"Check entry alerts:     [bold]investor watchlist-alerts[/bold]"
    )


# ---------------------------------------------------------------------------
# correction-scan command
# ---------------------------------------------------------------------------


@cli.command("correction-scan")
def correction_scan() -> None:
    """Check current market mode and show Tier-1 watchlist entry opportunities."""

    async def _run() -> tuple:
        from src.agent.mode_detector import detect_mode
        from src.api.nse import NSEClient
        from src.db.repository import get_watchlist_with_targets
        from src.models import AnalysisState

        nifty_state = AnalysisState(ticker="NIFTY")
        async with NSEClient() as nse_client:
            mode = await detect_mode(nse_client, nifty_state)
        tier1_rows = await get_watchlist_with_targets(settings.db_path)
        tier1_rows = [r for r in tier1_rows if (r.get("watchlist_tier") or 99) == 1]
        return mode, nifty_state, tier1_rows

    mode, nifty_state, tier1_rows = asyncio.run(_run())

    console.print(f"\n[bold]Market Mode: {mode.value.upper()}[/bold]")
    if nifty_state.nifty_level:
        console.print(f"Nifty 50: {nifty_state.nifty_level:,.2f}")
    if nifty_state.nifty_52w_high:
        console.print(f"52W High: {nifty_state.nifty_52w_high:,.2f}")
    if nifty_state.nifty_decline_pct:
        console.print(f"Decline from peak: {nifty_state.nifty_decline_pct:.2f}%")

    # Show Tier 1 watchlist from SQLite
    if tier1_rows:
        console.print("\n[bold cyan]Tier 1 Watchlist (from DB):[/bold cyan]")
        t1_table = Table(show_header=True, header_style="bold")
        t1_table.add_column("Ticker", style="bold", width=12)
        t1_table.add_column("Company", width=24)
        t1_table.add_column("Analysis Date", width=13)
        t1_table.add_column("CMP @ Analysis", justify="right", width=15)
        t1_table.add_column("Target Buy", justify="right", width=12)
        t1_table.add_column("Req MoS%", justify="right", width=10)
        for r in tier1_rows:
            cmp_val = r.get("cmp_at_analysis")
            target = r.get("target_buy_price")
            t1_table.add_row(
                r.get("ticker", ""),
                r.get("company_name", "") or "",
                r.get("analysis_date", ""),
                f"₹{cmp_val:.2f}" if cmp_val else "—",
                f"₹{target:.2f}" if target else "—",
                f"{r.get('required_mos_pct', '')}%",
            )
        console.print(t1_table)
        console.print(
            "[dim]Run [bold]investor watchlist-alerts[/bold] to compare live prices against targets.[/dim]"
        )
    else:
        console.print("\n[dim]No Tier 1 watchlist entries in database yet.[/dim]")


# ---------------------------------------------------------------------------
# scan command
# ---------------------------------------------------------------------------


@cli.command("scan")
@click.option(
    "--index",
    default="NIFTY 500",
    show_default=True,
    help="NSE index to scan (e.g. 'NIFTY 50', 'NIFTY 100', 'NIFTY 500')",
)
@click.option(
    "--top",
    default=10,
    show_default=True,
    help="Max number of full analyses to run after pre-screening",
)
@click.option(
    "--min-score",
    default=5,
    show_default=True,
    help="Minimum Step-0 pre-screen score (0–9) required to proceed to full analysis",
)
@click.option(
    "--prescreen-only",
    is_flag=True,
    default=False,
    help="Stop after Step-0 pre-screen; skip full 9-step analysis",
)
@click.option(
    "--concurrency",
    default=None,
    show_default=True,
    help=(
        "Max parallel HTTP requests during pre-screen. "
        "Defaults to settings.scan_concurrency (8). "
        "Use 3–5 on a cold scan; 10–15 once the SQLite warm cache is seeded."
    ),
)
def scan(
    index: str,
    top: int,
    min_score: int,
    prescreen_only: bool,
    concurrency: int | None,
) -> None:
    """Screen a stock universe and surface the best investment opportunities.

    Phase 1 — Pre-screen every stock in the index using deterministic Step 0
    (no Claude calls, no cost). Filters by minimum score.

    Phase 2 — Run the full 9-step pipeline on the shortlisted candidates.

    Examples:\n
      investor scan --index "NIFTY 50" --prescreen-only\n
      investor scan --index "NIFTY 100" --top 5 --min-score 7\n
      investor scan --concurrency 3
    """
    from src.agent.batch_scanner import BatchScanner

    async def _run():
        scanner = BatchScanner(concurrency=concurrency)
        return await scanner.scan(
            index=index,
            prescreen_min_score=min_score,
            max_full_analyses=top,
            prescreen_only=prescreen_only,
        )

    with console.status(f"[bold green]Scanning {index}...[/bold green]"):
        summaries, results = asyncio.run(_run())

    # --- Pre-screen table ---
    from rich.table import Table

    passed = [s for s in summaries if not s.error and s.score >= min_score]
    failed = [s for s in summaries if not s.error and s.score < min_score]
    errors = [s for s in summaries if s.error]

    ps_table = Table(
        title=f"Pre-Screen Results — {index}",
        show_header=True,
        header_style="bold cyan",
    )
    ps_table.add_column("Ticker", style="bold")
    ps_table.add_column("Score", justify="center")
    ps_table.add_column("Gate")
    ps_table.add_column("Failed Metrics", style="dim")

    gate_colours = {
        "pass_green": "green",
        "pass_conditional": "yellow",
        "fail": "red",
        "not_run": "dim",
    }

    for s in sorted(passed, key=lambda x: x.score, reverse=True):
        colour = gate_colours.get(s.gate.value, "white")
        ps_table.add_row(
            s.ticker,
            f"[{colour}]{s.score}/9[/{colour}]",
            f"[{colour}]{s.gate.value.upper()}[/{colour}]",
            ", ".join(s.failed_metrics[:3]) + ("…" if len(s.failed_metrics) > 3 else ""),
        )

    console.print(ps_table)
    console.print(
        f"\n[bold]Pre-screen summary:[/bold] "
        f"{len(passed)} passed / {len(failed)} failed / {len(errors)} errors "
        f"out of {len(summaries)} tickers"
    )

    if prescreen_only or not results:
        if not prescreen_only and passed:
            console.print(
                "\n[yellow]No full analyses completed "
                "(all candidates may have been rejected at Step 0).[/yellow]"
            )
        return

    # --- Full analysis table ---
    fa_table = Table(
        title="Full Analysis Results (ranked)",
        show_header=True,
        header_style="bold cyan",
    )
    fa_table.add_column("Ticker", style="bold")
    fa_table.add_column("Recommendation")
    fa_table.add_column("Conviction")
    fa_table.add_column("MoS %", justify="right")
    fa_table.add_column("Governance", justify="center")
    fa_table.add_column("Terminated At")

    rec_colours = {"BUY": "green", "WATCHLIST": "yellow", "PEER_SWITCH": "cyan", "REJECT": "red"}

    for state in results:
        rec = state.recommendation_type or "REJECT"
        colour = rec_colours.get(rec, "white")
        mos = (
            f"{state.valuation.margin_of_safety_pct:.1f}%"
            if state.valuation and state.valuation.margin_of_safety_pct is not None
            else "—"
        )
        gov = (
            f"{state.governance.score}/15"
            if state.governance
            else "—"
        )
        term = (
            f"Step {state.terminated_at_step}"
            if state.terminated_at_step is not None
            else "—"
        )
        fa_table.add_row(
            state.ticker,
            f"[{colour}]{rec}[/{colour}]",
            state.conviction.value.upper() if state.conviction else "—",
            mos,
            gov,
            term,
        )

    console.print(fa_table)

    # Print full report for BUY recommendations
    buy_states = [s for s in results if s.recommendation_type == "BUY"]
    if buy_states:
        console.print(f"\n[bold green]Found {len(buy_states)} BUY recommendation(s):[/bold green]")
        for state in buy_states:
            if state.formatted_output:
                console.print(state.formatted_output)
    else:
        console.print("\n[yellow]No BUY recommendations from this scan.[/yellow]")

    # Save reports and update watchlist/rejection tracker
    reports_dir = Path("analysis") / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    from datetime import date
    today = date.today().isoformat()
    for state in results:
        if state.formatted_output:
            path = reports_dir / f"{state.ticker}_{today}.md"
            path.write_text(state.formatted_output, encoding="utf-8")
    if results:
        console.print(f"[dim]Reports saved → {reports_dir}/[/dim]")


# ---------------------------------------------------------------------------
# portfolio-review command
# ---------------------------------------------------------------------------


@cli.command("portfolio-review")
@click.option(
    "--concurrency",
    default=2,
    show_default=True,
    help="Max parallel analyses (keep low — each uses Claude API)",
)
@click.option(
    "--save/--no-save",
    default=True,
    help="Save individual reports to analysis/reports/",
)
@click.option(
    "--user",
    "user_id",
    default=None,
    envvar="INVESTOR_USER",
    help="Portfolio user ID.",
)
def portfolio_review(concurrency: int, save: bool, user_id: str | None) -> None:
    """Run a full 9-step analysis on every holding and recommend HOLD or SELL.

    \b
    - BUY    → HOLD + consider adding on dips
    - WATCHLIST → HOLD with caution (set alert)
    - REJECT → SELL — thesis has broken down

    Reports are saved to analysis/reports/.
    """
    import time

    tracker = PortfolioTracker(user_id=user_id)

    async def _run_all():
        import asyncio as _aio

        holdings = await tracker.get_holdings()
        if not holdings:
            return holdings, []
        sem = _aio.Semaphore(concurrency)

        async def _one(ticker: str):
            async with sem:
                return await InvestmentPipeline().analyze(ticker)

        results = await _aio.gather(*[_one(h["ticker"]) for h in holdings], return_exceptions=True)
        return holdings, list(results)

    start = time.monotonic()
    with console.status("[bold green]Analysing portfolio...[/bold green]"):
        holdings, raw_results = asyncio.run(_run_all())
    elapsed = time.monotonic() - start

    if not holdings:
        console.print(
            f"[yellow]No holdings found for user '{tracker.user_id}'. "
            "Add positions with: investor add-trade TICKER BUY QTY PRICE[/yellow]"
        )
        return

    tickers = [h["ticker"] for h in holdings]
    holding_map = {h["ticker"]: h for h in holdings}

    console.print(
        f"[bold cyan]Portfolio Review — {len(tickers)} holdings[/bold cyan]\n"
        f"[dim]Running full 9-step analysis on: {', '.join(tickers)}[/dim]\n"
    )

    # Build result table
    action_table = Table(
        title="Portfolio Review — Hold / Sell Actions",
        show_header=True,
        header_style="bold cyan",
        show_lines=True,
    )
    action_table.add_column("Ticker", style="bold", width=12)
    action_table.add_column("Company", width=22)
    action_table.add_column("Avg Cost", justify="right", width=10)
    action_table.add_column("Qty", justify="right", width=6)
    action_table.add_column("Alloc %", justify="right", width=8)
    action_table.add_column("Pipeline", justify="center", width=12)
    action_table.add_column("Gov /15", justify="center", width=8)
    action_table.add_column("MoS %", justify="right", width=8)
    action_table.add_column("ACTION", style="bold", width=14)
    action_table.add_column("Reason", width=30)

    action_colours = {
        "HOLD_ADD": "green",
        "HOLD": "cyan",
        "HOLD_WATCH": "yellow",
        "SELL": "red",
    }

    holds, sells, errors = [], [], []

    for ticker, result in zip(tickers, raw_results):
        h = holding_map[ticker]

        if isinstance(result, Exception):
            errors.append(ticker)
            action_table.add_row(
                ticker,
                h.get("company_name", ""),
                f"₹{h['avg_cost']:.2f}",
                str(h["quantity"]),
                f"{h['allocation_pct']:.1f}%",
                "[red]ERROR[/red]",
                "—", "—",
                "[red]SKIP[/red]",
                str(result)[:30],
            )
            continue

        state: AnalysisState = result
        rec = state.recommendation_type or "REJECT"
        gov = f"{state.governance.score}/15" if state.governance else "—"
        mos = (
            f"{state.valuation.margin_of_safety_pct:.1f}%"
            if state.valuation and state.valuation.margin_of_safety_pct is not None
            else "—"
        )

        # Determine action
        if rec == "BUY":
            action = "HOLD_ADD"
            reason = "Thesis intact — add on dips"
            holds.append(ticker)
        elif rec == "WATCHLIST":
            action = "HOLD_WATCH"
            tier = state.watchlist_tier or 2
            reason = f"Tier {tier} — set re-entry alert"
            holds.append(ticker)
        else:
            action = "SELL"
            reason = (state.termination_reason or "Failed pipeline gate")[:30]
            sells.append(ticker)

        colour = action_colours.get(action, "white")
        action_table.add_row(
            ticker,
            h.get("company_name", ticker),
            f"₹{h['avg_cost']:.2f}",
            str(h["quantity"]),
            f"{h['allocation_pct']:.1f}%",
            f"[{'green' if rec == 'BUY' else 'yellow' if rec == 'WATCHLIST' else 'red'}]{rec}[/]",
            gov,
            mos,
            f"[{colour}]{action}[/{colour}]",
            reason,
        )

        # Save individual report
        if save and state.formatted_output:
            reports_dir = Path("analysis") / "reports"
            reports_dir.mkdir(parents=True, exist_ok=True)
            today = date.today().isoformat()
            path = reports_dir / f"{ticker}_{today}_portfolio_review.md"
            path.write_text(state.formatted_output, encoding="utf-8")

    console.print(action_table)

    # Count how many non-error results were saved (pipeline saves automatically)
    saved_count = sum(1 for r in raw_results if not isinstance(r, Exception))

    # Summary
    console.print(
        f"\n[bold]Summary:[/bold] "
        f"[green]{len(holds)} HOLD[/green]  |  "
        f"[red]{len(sells)} SELL[/red]"
        + (f"  |  [yellow]{len(errors)} ERROR[/yellow]" if errors else "")
        + f"  |  elapsed {elapsed:.0f}s"
    )
    console.print(
        f"[dim]Results saved to SQLite: {saved_count} analyses · {settings.db_path}[/dim]\n"
        "[dim]Use 'investor db-summary' to view all results · "
        "'investor db-history TICKER' for trend[/dim]"
    )

    if sells:
        console.print(
            f"\n[bold red]⚠ Sell candidates:[/bold red] {', '.join(sells)}\n"
            "Review individual reports in analysis/reports/ before acting.\n"
            "[dim]Tax note: check LTCG eligibility (>1Y = 12.5% on gains > ₹1.25L)[/dim]"
        )

    if errors:
        console.print(
            f"\n[yellow]Could not analyse: {', '.join(errors)}. "
            "Re-run individually with: investor analyze TICKER[/yellow]"
        )


# ---------------------------------------------------------------------------
# db-history command
# ---------------------------------------------------------------------------


@cli.command("db-history")
@click.argument("ticker")
@click.option("--limit", default=10, show_default=True, help="Number of past analyses to show")
def db_history(ticker: str, limit: int) -> None:
    """Show analysis history for TICKER from the local SQLite database.

    Example: investor db-history RELIANCE --limit 5
    """
    import asyncio as _asyncio

    from src.db.repository import get_analysis_history

    rows = _asyncio.run(get_analysis_history(settings.db_path, ticker.upper(), limit))
    if not rows:
        console.print(f"[yellow]No history found for {ticker.upper()} in {settings.db_path}[/yellow]")
        return

    table = Table(
        title=f"Analysis History — {ticker.upper()}",
        show_header=True,
        header_style="bold cyan",
    )
    table.add_column("Date")
    table.add_column("Recommendation")
    table.add_column("Conviction")
    table.add_column("Sector")
    table.add_column("Gov /15", justify="center")
    table.add_column("Fin /7", justify="center")
    table.add_column("MoS %", justify="right")
    table.add_column("Terminated At")

    rec_colours = {"BUY": "green", "WATCHLIST": "yellow", "PEER_SWITCH": "cyan", "REJECT": "red"}

    for row in rows:
        rec = row.get("recommendation") or "—"
        colour = rec_colours.get(rec, "white")
        mos = f"{row['mos_pct']:.1f}%" if row.get("mos_pct") is not None else "—"
        term = f"Step {row['terminated_at_step']}" if row.get("terminated_at_step") is not None else "—"
        table.add_row(
            row.get("analysis_date", ""),
            f"[{colour}]{rec}[/{colour}]",
            row.get("conviction") or "—",
            row.get("sector_name") or "—",
            str(row.get("governance_score") or "—"),
            str(row.get("financial_score") or "—"),
            mos,
            term,
        )

    console.print(table)


# ---------------------------------------------------------------------------
# db-summary command
# ---------------------------------------------------------------------------


@cli.command("db-summary")
def db_summary() -> None:
    """Show a summary of all analyses in the local SQLite database.

    Displays all analysed tickers ranked by recommendation type.
    """
    import asyncio as _asyncio

    from src.db.repository import get_summary

    rows = _asyncio.run(get_summary(settings.db_path))
    if not rows:
        console.print(f"[yellow]No analyses found in {settings.db_path}[/yellow]")
        return

    table = Table(
        title="Database Summary — All Analyses",
        show_header=True,
        header_style="bold cyan",
    )
    table.add_column("Ticker", style="bold")
    table.add_column("Company")
    table.add_column("Date")
    table.add_column("Recommendation")
    table.add_column("Conviction")
    table.add_column("Sector")
    table.add_column("Cap")
    table.add_column("MoS %", justify="right")
    table.add_column("Terminated At")

    rec_colours = {"BUY": "green", "WATCHLIST": "yellow", "PEER_SWITCH": "cyan", "REJECT": "red"}

    for row in rows:
        rec = row.get("recommendation") or "—"
        colour = rec_colours.get(rec, "white")
        mos = f"{row['mos_pct']:.1f}%" if row.get("mos_pct") is not None else "—"
        term = f"Step {row['terminated_at_step']}" if row.get("terminated_at_step") is not None else "—"
        table.add_row(
            row.get("ticker", ""),
            (row.get("company_name") or "")[:25],
            row.get("analysis_date", ""),
            f"[{colour}]{rec}[/{colour}]",
            row.get("conviction") or "—",
            row.get("sector_name") or "—",
            row.get("cap_size") or "—",
            mos,
            term,
        )

    console.print(table)
    console.print(f"\n[dim]Total: {len(rows)} analyses in {settings.db_path}[/dim]")


# ---------------------------------------------------------------------------
# db-recommendations command
# ---------------------------------------------------------------------------


@cli.command("db-recommendations")
@click.option(
    "--type",
    "rec_type",
    default="BUY",
    type=click.Choice(["BUY", "WATCHLIST", "PEER_SWITCH", "REJECT"], case_sensitive=False),
    show_default=True,
    help="Recommendation type to filter",
)
def db_recommendations(rec_type: str) -> None:
    """List all stocks with a given recommendation from the SQLite database.

    \b
    Examples:
      investor db-recommendations --type BUY
      investor db-recommendations --type WATCHLIST
    """
    import asyncio as _asyncio

    from src.db.repository import list_recommendations

    rows = _asyncio.run(list_recommendations(settings.db_path, rec_type.upper()))
    if not rows:
        console.print(f"[yellow]No {rec_type.upper()} recommendations in {settings.db_path}[/yellow]")
        return

    rec_colours = {"BUY": "green", "WATCHLIST": "yellow", "PEER_SWITCH": "cyan", "REJECT": "red"}
    colour = rec_colours.get(rec_type.upper(), "white")

    table = Table(
        title=f"[{colour}]{rec_type.upper()}[/{colour}] Recommendations",
        show_header=True,
        header_style="bold cyan",
        show_lines=False,
    )
    table.add_column("Ticker", style="bold", width=12)
    table.add_column("Company", width=28)
    table.add_column("Date", width=12)
    table.add_column("Conviction", width=10)
    table.add_column("Sector", width=20)
    table.add_column("Cap", width=10)
    table.add_column("Gov /15", justify="center", width=8)
    table.add_column("Fin /7", justify="center", width=7)
    table.add_column("MoS %", justify="right", width=8)
    table.add_column("DCF ₹", justify="right", width=10)

    for row in rows:
        mos = f"{row['mos_pct']:.1f}%" if row.get("mos_pct") is not None else "—"
        dcf = f"₹{row['dcf_intrinsic_weighted']:.0f}" if row.get("dcf_intrinsic_weighted") else "—"
        table.add_row(
            row.get("ticker", ""),
            (row.get("company_name") or "")[:28],
            row.get("analysis_date", ""),
            row.get("conviction") or "—",
            row.get("sector_name") or "—",
            row.get("cap_size") or "—",
            str(row.get("governance_score") or "—"),
            str(row.get("financial_score") or "—"),
            mos,
            dcf,
        )

    console.print(table)
    console.print(f"\n[dim]{len(rows)} stocks · {settings.db_path}[/dim]")


# ---------------------------------------------------------------------------
# db-snapshots command
# ---------------------------------------------------------------------------


@cli.command("db-snapshots")
@click.argument("ticker")
@click.option(
    "--type",
    "data_type",
    default=None,
    type=click.Choice(["quote", "financials", "governance", "valuation"]),
    help="Filter by data type (default: show all)",
)
def db_snapshots(ticker: str, data_type: str | None) -> None:
    """Show raw data snapshots stored for TICKER in the SQLite database.

    \b
    Examples:
      investor db-snapshots RELIANCE
      investor db-snapshots RELIANCE --type financials
    """
    import asyncio as _asyncio
    import json

    import aiosqlite

    ticker = ticker.upper()

    async def _fetch():
        from src.db.repository import init_db as _init_db
        await _init_db(settings.db_path)
        async with aiosqlite.connect(settings.db_path) as db:
            db.row_factory = aiosqlite.Row
            if data_type:
                sql = (
                    "SELECT snapshot_date, data_type, source, data_json FROM data_snapshots "
                    "WHERE ticker = ? AND data_type = ? ORDER BY snapshot_date DESC, data_type"
                )
                params = (ticker, data_type)
            else:
                sql = (
                    "SELECT snapshot_date, data_type, source, data_json FROM data_snapshots "
                    "WHERE ticker = ? ORDER BY snapshot_date DESC, data_type"
                )
                params = (ticker,)
            async with db.execute(sql, params) as cursor:
                return [dict(r) async for r in cursor]

    rows = _asyncio.run(_fetch())

    if not rows:
        msg = f"[yellow]No snapshots for {ticker}"
        if data_type:
            msg += f" ({data_type})"
        console.print(msg + f" in {settings.db_path}[/yellow]")
        return

    table = Table(
        title=f"Data Snapshots — {ticker}",
        show_header=True,
        header_style="bold cyan",
        show_lines=True,
    )
    table.add_column("Date", width=12)
    table.add_column("Type", width=12)
    table.add_column("Source", width=12)
    table.add_column("Fields (preview)", width=60)

    for row in rows:
        try:
            data = json.loads(row["data_json"])
            # Show only a compact preview of the stored fields
            keys = list(data.keys())
            preview_keys = keys[:6]
            preview = ", ".join(
                f"{k}={data[k]!r}" for k in preview_keys if data[k] is not None
            )
            if len(keys) > 6:
                preview += f"  … +{len(keys) - 6} more"
        except Exception:
            preview = row["data_json"][:80]

        table.add_row(
            row.get("snapshot_date", ""),
            row.get("data_type", ""),
            row.get("source", ""),
            preview,
        )

    console.print(table)
    console.print(f"\n[dim]{len(rows)} snapshot(s) for {ticker} · {settings.db_path}[/dim]")


# ---------------------------------------------------------------------------
# watchlist-alerts command  (P2-2)
# ---------------------------------------------------------------------------


@cli.command("watchlist-alerts")
def watchlist_alerts() -> None:
    """Compare every WATCHLIST ticker's DCF target buy price against live CMP.

    Fetches current prices via Yahoo Finance and flags tickers that have
    entered (or are close to) their required margin-of-safety zone.

    \b
    Alert levels:
      🟢 ENTER ZONE   — CMP ≤ target buy price (MoS met — run full analysis now)
      🟡 APPROACHING  — CMP within 10 % above target
      ⚪ MONITORING   — CMP still above target zone
    """
    import asyncio as _asyncio

    from src.db.repository import get_watchlist_with_targets

    rows = _asyncio.run(get_watchlist_with_targets(settings.db_path))

    if not rows:
        console.print(
            f"[yellow]No WATCHLIST tickers found in {settings.db_path}. "
            "Run 'investor analyze TICKER' to add stocks to the watchlist.[/yellow]"
        )
        return

    async def _fetch_prices(tickers: list[str]) -> dict[str, float | None]:
        """Fetch live CMP for a list of tickers via YFinance."""
        from src.api.yfinance_client import YFinanceClient

        async with YFinanceClient() as yf_client:
            results: dict[str, float | None] = {}
            for t in tickers:
                quote = await yf_client.get_stock_quote(t)
                results[t] = quote.cmp if quote else None
        return results

    tickers = [r["ticker"] for r in rows]
    with console.status("[bold green]Fetching live prices...[/bold green]"):
        live_prices = asyncio.run(_fetch_prices(tickers))

    table = Table(
        title="Watchlist Alerts — Live CMP vs DCF Target",
        show_header=True,
        header_style="bold cyan",
        show_lines=True,
    )
    table.add_column("Ticker", style="bold", width=10)
    table.add_column("Company", width=22)
    table.add_column("Tier", justify="center", width=5)
    table.add_column("Last Analysis", width=12)
    table.add_column("CMP @ Analysis", justify="right", width=14)
    table.add_column("Live CMP", justify="right", width=10)
    table.add_column("Target Buy ₹", justify="right", width=12)
    table.add_column("Gap %", justify="right", width=8)
    table.add_column("Status", width=16)

    enter_zone, approaching, monitoring = [], [], []

    for row in rows:
        ticker = row["ticker"]
        live_cmp = live_prices.get(ticker)
        target = row.get("target_buy_price")
        cmp_at_analysis = row.get("cmp_at_analysis")
        tier = row.get("watchlist_tier") or "—"

        live_str = f"₹{live_cmp:.2f}" if live_cmp else "[dim]N/A[/dim]"
        target_str = f"₹{target:.2f}" if target else "—"
        old_cmp_str = f"₹{cmp_at_analysis:.2f}" if cmp_at_analysis else "—"

        if live_cmp and target:
            gap_pct = (live_cmp - target) / target * 100
            gap_str = f"{gap_pct:+.1f}%"
            if gap_pct <= 0:
                status = "[green]🟢 ENTER ZONE[/green]"
                enter_zone.append(ticker)
            elif gap_pct <= 10:
                status = "[yellow]🟡 APPROACHING[/yellow]"
                approaching.append(ticker)
            else:
                status = "[dim]⚪ MONITORING[/dim]"
                monitoring.append(ticker)
        else:
            gap_str = "—"
            status = "[dim]⚪ NO TARGET[/dim]"
            monitoring.append(ticker)

        table.add_row(
            ticker,
            (row.get("company_name") or "")[:22],
            str(tier),
            row.get("analysis_date", ""),
            old_cmp_str,
            live_str,
            target_str,
            gap_str,
            status,
        )

    console.print(table)

    summary_parts = []
    if enter_zone:
        summary_parts.append(f"[green]{len(enter_zone)} in buy zone: {', '.join(enter_zone)}[/green]")
    if approaching:
        summary_parts.append(f"[yellow]{len(approaching)} approaching: {', '.join(approaching)}[/yellow]")
    if monitoring:
        summary_parts.append(f"[dim]{len(monitoring)} monitoring[/dim]")
    console.print("\n" + "  |  ".join(summary_parts))

    if enter_zone:
        console.print(
            f"\n[bold green]⚡ Action required:[/bold green] "
            f"Run full analysis on: {', '.join(f'investor analyze {t}' for t in enter_zone)}"
        )
    console.print(
        "\n[dim]Prices via Yahoo Finance (~15-20 min delayed). "
        "Targets derived from DCF intrinsic at time of last analysis.[/dim]"
    )


# ---------------------------------------------------------------------------
# surveillance command  (P2-1)
# ---------------------------------------------------------------------------


@cli.command("surveillance")
@click.option(
    "--days-since",
    default=30,
    show_default=True,
    help="Flag tickers whose last analysis is older than this many days",
)
def surveillance(days_since: int) -> None:
    """Light-touch surveillance sweep across all tracked BUY and WATCHLIST positions.

    For each ticker in the database with a BUY or WATCHLIST recommendation:
      1. Fetches live CMP via Yahoo Finance
      2. Checks price drift since last analysis (big drops may mean thesis change)
      3. Flags stale analyses (> --days-since days old)
      4. Highlights watchlist tickers that have entered their buy zone

    This is a quick, zero-LLM scan (~5s for 30 tickers).  When it flags
    a ticker for re-analysis, run: investor analyze TICKER

    \b
    Drift alerts:
      ▲ CMP up > 20 % since analysis  → may no longer offer MoS; re-evaluate
      ▼ CMP down > 20 % since analysis → possible buying opportunity or thesis break
    """
    import asyncio as _asyncio
    from datetime import datetime as _dt

    from src.db.repository import get_all_tracked_tickers

    rows = _asyncio.run(get_all_tracked_tickers(settings.db_path))

    if not rows:
        console.print(
            f"[yellow]No BUY or WATCHLIST tickers found in {settings.db_path}.[/yellow]"
        )
        return

    async def _fetch_prices(tickers: list[str]) -> dict[str, float | None]:
        from src.api.yfinance_client import YFinanceClient

        async with YFinanceClient() as yf_client:
            out: dict[str, float | None] = {}
            for t in tickers:
                quote = await yf_client.get_stock_quote(t)
                out[t] = quote.cmp if quote else None
        return out

    tickers = [r["ticker"] for r in rows]
    with console.status("[bold green]Fetching live prices for surveillance...[/bold green]"):
        live_prices = asyncio.run(_fetch_prices(tickers))

    today = date.today()

    table = Table(
        title="Surveillance — All BUY / WATCHLIST Positions",
        show_header=True,
        header_style="bold cyan",
        show_lines=False,
    )
    table.add_column("Ticker", style="bold", width=10)
    table.add_column("Rec", width=10)
    table.add_column("Last Analysis", width=12)
    table.add_column("Stale?", justify="center", width=8)
    table.add_column("CMP @ Analysis", justify="right", width=14)
    table.add_column("Live CMP", justify="right", width=10)
    table.add_column("Drift %", justify="right", width=9)
    table.add_column("Target ₹", justify="right", width=10)
    table.add_column("Zone?", justify="center", width=8)
    table.add_column("Action", width=24)

    re_analyse = []
    enter_zone = []

    for row in rows:
        ticker = row["ticker"]
        rec = row.get("recommendation", "")
        analysis_date_str = row.get("analysis_date", "")
        cmp_prev = row.get("cmp_at_analysis")
        target = row.get("target_buy_price")
        live_cmp = live_prices.get(ticker)

        # Staleness
        try:
            analysis_date = _dt.fromisoformat(analysis_date_str).date()
            age_days = (today - analysis_date).days
            stale = age_days > days_since
        except Exception:
            age_days = 999
            stale = True
        stale_str = f"[red]{age_days}d[/red]" if stale else f"[green]{age_days}d[/green]"

        # Price drift
        drift_str = "—"
        drift_action = ""
        if live_cmp and cmp_prev and cmp_prev > 0:
            drift_pct = (live_cmp - cmp_prev) / cmp_prev * 100
            drift_str = f"{drift_pct:+.1f}%"
            if drift_pct > 20:
                drift_str = f"[yellow]{drift_str}[/yellow]"
                drift_action = "↑ Re-evaluate MoS"
            elif drift_pct < -20:
                drift_str = f"[cyan]{drift_str}[/cyan]"
                drift_action = "↓ Opportunity / thesis?"

        # Watchlist zone check
        zone_str = "—"
        if rec == "WATCHLIST" and live_cmp and target:
            gap = (live_cmp - target) / target * 100
            if gap <= 0:
                zone_str = "[green]✓ IN[/green]"
                enter_zone.append(ticker)
            elif gap <= 10:
                zone_str = "[yellow]~10%[/yellow]"
            else:
                zone_str = f"[dim]+{gap:.0f}%[/dim]"

        # Recommended action
        if stale:
            action = "[yellow]Re-run analysis (stale)[/yellow]"
            re_analyse.append(ticker)
        elif drift_action:
            action = drift_action
            re_analyse.append(ticker)
        elif zone_str.startswith("[green]"):
            action = "[green]Run full analysis now[/green]"
        else:
            action = "[dim]Continue monitoring[/dim]"

        rec_colour = "green" if rec == "BUY" else "yellow"
        live_str = f"₹{live_cmp:.2f}" if live_cmp else "[dim]N/A[/dim]"
        prev_str = f"₹{cmp_prev:.2f}" if cmp_prev else "—"
        target_str = f"₹{target:.2f}" if target else "—"

        table.add_row(
            ticker,
            f"[{rec_colour}]{rec}[/{rec_colour}]",
            analysis_date_str,
            stale_str,
            prev_str,
            live_str,
            drift_str,
            target_str,
            zone_str,
            action,
        )

    console.print(table)

    # Summary actions
    if enter_zone:
        console.print(
            f"\n[bold green]🟢 {len(enter_zone)} ticker(s) in buy zone:[/bold green] "
            + ", ".join(enter_zone)
        )
    if re_analyse:
        unique_reanalyse = list(dict.fromkeys(re_analyse))  # dedupe, preserve order
        console.print(
            f"\n[bold yellow]⚡ {len(unique_reanalyse)} ticker(s) need re-analysis:[/bold yellow] "
            + ", ".join(f"[bold]{t}[/bold]" for t in unique_reanalyse[:10])
        )
        console.print(
            "[dim]Run: " + "  |  ".join(f"investor analyze {t}" for t in unique_reanalyse[:5])
            + ("[dim]  …" if len(unique_reanalyse) > 5 else "") + "[/dim]"
        )

    console.print(
        f"\n[dim]Prices via Yahoo Finance (~15-20 min delayed). "
        f"Stale threshold: {days_since} days. "
        f"DB: {settings.db_path}[/dim]"
    )


# ---------------------------------------------------------------------------
# import-pnl command
# ---------------------------------------------------------------------------


@cli.command("import-pnl")
@click.argument("xlsx_file", type=click.Path(exists=True, dir_okay=False))
@click.option("--user", "user_id", default=None, envvar="INVESTOR_USER", help="Portfolio user ID.")
@click.option("--dry-run", is_flag=True, default=False, help="Preview without writing to DB.")
@click.option("--clear", is_flag=True, default=False, help="Clear existing portfolio data first.")
def import_pnl(xlsx_file: str, user_id: str | None, dry_run: bool, clear: bool) -> None:
    """Import holdings and BUY transactions from a Zerodha P&L Excel report.

    Extracts two things from the Equity sheet:
      - Open positions  → portfolio_holdings  (avg_cost = open_value / open_qty)
      - Traded stocks   → BUY transactions    (avg_price = buy_value / quantity)

    Sell transactions should be imported separately via import-tradebook.

    \b
    Example:
      investor import-pnl ~/Downloads/pnl-UTT486.xlsx --user rm --clear
    """
    import openpyxl

    tracker = PortfolioTracker(user_id=user_id)

    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    ws = wb["Equity"]

    # Locate header row (contains 'Symbol')
    header_row = None
    for row in ws.iter_rows(values_only=True):
        if row[1] == "Symbol":
            header_row = row
            break
    if header_row is None:
        console.print("[red]Could not find header row in Equity sheet.[/red]")
        raise SystemExit(1)

    # Parse period from sheet for use as BUY date placeholder
    period_start = date(2025, 6, 1)
    for row in ws.iter_rows(values_only=True):
        if row[1] and isinstance(row[1], str) and "P&L Statement" in row[1]:
            import re as _re
            m = _re.search(r"from (\d{4}-\d{2}-\d{2})", row[1])
            if m:
                period_start = date.fromisoformat(m.group(1))
            break

    holdings, buy_txns = [], []
    past_header = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[1] == "Symbol":
            past_header = True
            continue
        if not past_header:
            continue
        symbol = row[1]
        if not symbol or not isinstance(symbol, str) or symbol.strip() == "":
            continue
        try:
            qty        = float(row[3] or 0)   # col D: Quantity (traded)
            buy_val    = float(row[4] or 0)   # col E: Buy Value
            open_qty   = float(row[9] or 0) if row[9] not in ("", None) else 0.0   # col J
            open_val   = float(row[11] or 0) if row[11] not in ("", None) else 0.0  # col L
        except (TypeError, ValueError):
            continue

        ticker = symbol.upper().strip()

        # Open position → holding
        if open_qty > 0 and open_val > 0:
            avg_cost = round(open_val / open_qty, 4)
            holdings.append({
                "ticker": ticker,
                "qty": int(open_qty),
                "avg_cost": avg_cost,
                "open_val": open_val,
            })

        # Traded position → BUY transaction
        if qty > 0 and buy_val > 0:
            avg_buy = round(buy_val / qty, 4)
            buy_txns.append({
                "ticker": ticker,
                "qty": int(qty),
                "avg_buy": avg_buy,
                "buy_val": buy_val,
            })

    # ── Preview tables ────────────────────────────────────────────────────
    h_table = Table(
        title=f"Holdings (open positions) — {len(holdings)} stocks",
        header_style="bold cyan",
    )
    h_table.add_column("Ticker", style="bold", width=16)
    h_table.add_column("Qty", justify="right", width=6)
    h_table.add_column("Avg Cost ₹", justify="right", width=12)
    h_table.add_column("Value ₹", justify="right", width=12)
    for h in sorted(holdings, key=lambda x: x["ticker"]):
        h_table.add_row(h["ticker"], str(h["qty"]), f"{h['avg_cost']:,.2f}", f"{h['open_val']:,.0f}")
    console.print(h_table)

    b_table = Table(
        title=f"BUY Transactions — {len(buy_txns)} stocks",
        header_style="bold green",
    )
    b_table.add_column("Ticker", style="bold", width=16)
    b_table.add_column("Qty", justify="right", width=6)
    b_table.add_column("Avg Buy ₹", justify="right", width=12)
    b_table.add_column("Buy Value ₹", justify="right", width=12)
    for t in sorted(buy_txns, key=lambda x: x["ticker"]):
        b_table.add_row(t["ticker"], str(t["qty"]), f"{t['avg_buy']:,.2f}", f"{t['buy_val']:,.0f}")
    console.print(b_table)
    console.print(f"[dim]BUY date placeholder: {period_start} (P&L period start)[/dim]")

    if dry_run:
        console.print("[yellow]Dry run — nothing written to DB.[/yellow]")
        return

    # ── Clear if requested ────────────────────────────────────────────────
    if clear:
        from src.db.repository import clear_portfolio
        counts = asyncio.run(clear_portfolio(settings.db_path, tracker.user_id))
        console.print(f"[yellow]Cleared {sum(counts.values())} rows for user [bold]{tracker.user_id}[/bold][/yellow]")

    # ── Insert ────────────────────────────────────────────────────────────
    async def _insert() -> None:
        for h in holdings:
            await tracker.add_holding(
                ticker=h["ticker"],
                avg_cost=h["avg_cost"],
                quantity=h["qty"],
                purchase_date=period_start,
                allocation_pct=0.0,
                company_name=h["ticker"],
            )
        for t in buy_txns:
            await tracker.add_transaction(
                ticker=t["ticker"],
                action="BUY",
                price=t["avg_buy"],
                quantity=t["qty"],
                txn_date=period_start,
                notes=f"P&L import (buy_val={t['buy_val']:.2f})",
            )

    asyncio.run(_insert())
    console.print(
        f"\n[green]✓[/green] {len(holdings)} holdings + {len(buy_txns)} BUY transactions "
        f"imported for user [bold]{tracker.user_id}[/bold] → {settings.db_path}"
    )


# ---------------------------------------------------------------------------
# import-tradebook command
# ---------------------------------------------------------------------------


@cli.command("import-tradebook")
@click.argument("csv_file", type=click.Path(exists=True, dir_okay=False))
@click.option(
    "--user",
    "user_id",
    default=None,
    envvar="INVESTOR_USER",
    help="Portfolio user ID.",
)
@click.option(
    "--dry-run",
    is_flag=True,
    default=False,
    help="Preview aggregated transactions without writing to DB.",
)
@click.option(
    "--clear",
    is_flag=True,
    default=False,
    help="Delete all existing portfolio data for this user before importing.",
)
def import_tradebook(csv_file: str, user_id: str | None, dry_run: bool, clear: bool) -> None:
    """Import a Zerodha tradebook CSV into the portfolio database.

    Zerodha splits large orders across multiple fill rows sharing the same
    order_id.  This command aggregates fills into one logical transaction per
    order (weighted-average price, summed quantity) before inserting.

    \b
    Example:
      investor import-tradebook ~/Downloads/tradebook-UTT486-EQ.csv --user rm
      investor import-tradebook ~/Downloads/tradebook.csv --user rm --dry-run
    """
    import csv
    from collections import defaultdict

    tracker = PortfolioTracker(user_id=user_id)

    # ── 1. Parse CSV and group fills by order_id ──────────────────────────
    orders: dict[str, list[dict]] = defaultdict(list)
    with open(csv_file, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            orders[row["order_id"]].append(row)

    # ── 2. Aggregate fills into one logical transaction per order ─────────
    txns = []
    for order_id, fills in orders.items():
        total_qty = sum(float(f["quantity"]) for f in fills)
        total_value = sum(float(f["quantity"]) * float(f["price"]) for f in fills)
        avg_price = total_value / total_qty if total_qty else 0.0

        first = fills[0]
        symbol = first["symbol"].upper().strip()
        action = first["trade_type"].upper()  # BUY / SELL
        trade_date = first["trade_date"]      # YYYY-MM-DD

        txns.append({
            "order_id": order_id,
            "ticker": symbol,
            "action": action,
            "qty": int(total_qty),
            "avg_price": round(avg_price, 4),
            "date": trade_date,
            "fills": len(fills),
        })

    txns.sort(key=lambda t: (t["date"], t["ticker"]))

    # ── 3. Preview table ──────────────────────────────────────────────────
    table = Table(
        title=f"Tradebook — {len(txns)} orders from {csv_file}",
        show_header=True,
        header_style="bold cyan",
    )
    table.add_column("Date", width=12)
    table.add_column("Ticker", style="bold", width=16)
    table.add_column("Action", width=6)
    table.add_column("Qty", justify="right", width=6)
    table.add_column("Avg Price ₹", justify="right", width=12)
    table.add_column("Value ₹", justify="right", width=12)
    table.add_column("Fills", justify="right", width=5)

    for t in txns:
        colour = "green" if t["action"] == "BUY" else "red"
        table.add_row(
            t["date"],
            t["ticker"],
            f"[{colour}]{t['action']}[/{colour}]",
            str(t["qty"]),
            f"{t['avg_price']:,.2f}",
            f"{t['qty'] * t['avg_price']:,.0f}",
            str(t["fills"]),
        )

    console.print(table)

    if dry_run:
        console.print("[yellow]Dry run — nothing written to DB.[/yellow]")
        return

    # ── 4. Clear existing data if requested ───────────────────────────────
    if clear:
        from src.db.repository import clear_portfolio
        counts = asyncio.run(clear_portfolio(settings.db_path, tracker.user_id))
        total_cleared = sum(counts.values())
        console.print(
            f"[yellow]Cleared {total_cleared} existing rows for user "
            f"[bold]{tracker.user_id}[/bold][/yellow]"
        )

    # ── 5. Insert ─────────────────────────────────────────────────────────
    async def _insert() -> None:
        for t in txns:
            await tracker.add_transaction(
                ticker=t["ticker"],
                action=t["action"],
                price=t["avg_price"],
                quantity=t["qty"],
                txn_date=date.fromisoformat(t["date"]),
                notes=f"Zerodha order {t['order_id']} ({t['fills']} fill{'s' if t['fills'] > 1 else ''})",
            )

    asyncio.run(_insert())
    console.print(
        f"\n[green]✓[/green] {len(txns)} transactions imported for user "
        f"[bold]{tracker.user_id}[/bold] → {settings.db_path}"
    )


# ---------------------------------------------------------------------------
# clear-portfolio command
# ---------------------------------------------------------------------------


@cli.command("clear-portfolio")
@click.option(
    "--user",
    "user_id",
    default=None,
    envvar="INVESTOR_USER",
    help="Portfolio user ID whose data will be deleted.",
)
@click.confirmation_option(prompt="Delete ALL portfolio data for this user?")
def clear_portfolio_cmd(user_id: str | None) -> None:
    """Delete all holdings, transactions, and tax records for a user.

    \b
    Example:
      investor clear-portfolio --user rm
    """
    from src.db.repository import clear_portfolio

    uid = user_id or settings.investor_user
    counts = asyncio.run(clear_portfolio(settings.db_path, uid))
    total = sum(counts.values())
    console.print(
        f"[green]✓[/green] Cleared {total} rows for user [bold]{uid}[/bold] "
        f"(holdings={counts['portfolio_holdings']}, "
        f"transactions={counts['portfolio_transactions']}, "
        f"tax={counts['portfolio_tax']})"
    )


# ---------------------------------------------------------------------------
# migrate-portfolio command
# ---------------------------------------------------------------------------


@cli.command("migrate-portfolio")
@click.option(
    "--user",
    "user_id",
    required=True,
    envvar="INVESTOR_USER",
    help="User ID to assign all migrated records (e.g. 'rm').",
)
@click.option(
    "--portfolio-dir",
    default="portfolio",
    show_default=True,
    help="Directory containing the legacy markdown files.",
)
def migrate_portfolio(user_id: str, portfolio_dir: str) -> None:
    """Migrate legacy portfolio markdown files into the SQLite database.

    Reads holdings.md, transaction-log.md, and tax-tracker.md from
    --portfolio-dir and inserts every row into the DB under --user.
    Safe to re-run: duplicate rows are inserted as additional records
    (not upserted), so run only once per file.

    \b
    Example:
      investor migrate-portfolio --user rm
    """
    from pathlib import Path

    import aiosqlite

    from src.db.repository import init_db

    pdir = Path(portfolio_dir)

    async def _migrate() -> dict[str, int]:
        await init_db(settings.db_path)
        counts: dict[str, int] = {"holdings": 0, "transactions": 0, "tax": 0}

        # --- holdings.md ---
        hfile = pdir / "holdings.md"
        if hfile.exists():
            async with aiosqlite.connect(settings.db_path) as db:
                for line in hfile.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line.startswith("|") or "---" in line or "Ticker" in line:
                        continue
                    cells = [c.strip() for c in line.strip("|").split("|")]
                    if len(cells) < 6:
                        continue
                    try:
                        await db.execute(
                            "INSERT INTO portfolio_holdings "
                            "(user_id,ticker,company_name,avg_cost,quantity,purchase_date,allocation_pct) "
                            "VALUES (?,?,?,?,?,?,?)",
                            (
                                user_id,
                                cells[0].upper(),
                                cells[1],
                                float(cells[2].replace("₹", "").replace(",", "")),
                                int(cells[3]),
                                cells[4],
                                float(cells[5].replace("%", "").strip()),
                            ),
                        )
                        counts["holdings"] += 1
                    except (ValueError, IndexError):
                        pass
                await db.commit()

        # --- transaction-log.md ---
        tfile = pdir / "transaction-log.md"
        if tfile.exists():
            async with aiosqlite.connect(settings.db_path) as db:
                for line in tfile.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line.startswith("|") or "---" in line or "Date" in line:
                        continue
                    cells = [c.strip() for c in line.strip("|").split("|")]
                    if len(cells) < 5:
                        continue
                    try:
                        await db.execute(
                            "INSERT INTO portfolio_transactions "
                            "(user_id,txn_date,ticker,action,price,quantity,notes) "
                            "VALUES (?,?,?,?,?,?,?)",
                            (
                                user_id,
                                cells[0],
                                cells[1].upper(),
                                cells[2].upper(),
                                float(cells[3].replace("₹", "").replace(",", "")),
                                int(cells[4]),
                                cells[5] if len(cells) > 5 else "",
                            ),
                        )
                        counts["transactions"] += 1
                    except (ValueError, IndexError):
                        pass
                await db.commit()

        # --- tax-tracker.md ---
        xfile = pdir / "tax-tracker.md"
        if xfile.exists():
            async with aiosqlite.connect(settings.db_path) as db:
                for line in xfile.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line.startswith("|") or "---" in line or "Ticker" in line:
                        continue
                    cells = [c.strip() for c in line.strip("|").split("|")]
                    if len(cells) < 3:
                        continue
                    try:
                        await db.execute(
                            "INSERT INTO portfolio_tax "
                            "(user_id,ticker,purchase_date,ltcg_date,avg_cost) "
                            "VALUES (?,?,?,?,?)",
                            (
                                user_id,
                                cells[0].upper(),
                                cells[1],
                                cells[2],
                                float(cells[3].replace("₹", "").replace(",", "")) if len(cells) > 3 else 0.0,
                            ),
                        )
                        counts["tax"] += 1
                    except (ValueError, IndexError):
                        pass
                await db.commit()

        return counts

    counts = asyncio.run(_migrate())
    console.print(f"[green]✓[/green] Migration complete for user [bold]{user_id}[/bold]:")
    console.print(f"  Holdings:     {counts['holdings']} rows")
    console.print(f"  Transactions: {counts['transactions']} rows")
    console.print(f"  Tax entries:  {counts['tax']} rows")
    if all(v == 0 for v in counts.values()):
        console.print(
            f"[dim]No markdown files found in '{portfolio_dir}/' or all files were empty.[/dim]"
        )


if __name__ == "__main__":
    cli()
